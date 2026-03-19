from flask import Flask, jsonify, send_from_directory, request
from flask_cors import CORS
from db import get_connection
import os
from docxtpl import DocxTemplate
from flask import send_file
from io import BytesIO
from datetime import datetime, timedelta
from calendar import monthrange

BASE_DIR = os.path.dirname(os.path.dirname(__file__))

app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, "frontend"),
    static_url_path=""
)

CORS(app)


# SERVIR FRONTEND
@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


# REGISTRAR BENEFICIARIO
@app.route("/api/beneficiarios", methods=["POST"])
def registrar_beneficiario():
    data = request.json
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO beneficiarios
    (ci,nombre,cite,nudecud,tipo_codigo,fecha_inicio,carga_horaria,periodo_meses,telefono)
    VALUES (?,?,?,?,?,?,?,?,?)
    """,(
        data["ci"], data["nombre"], data["cite"], data["nudecud"],
        data["tipo_codigo"], data["fecha_inicio"], data["carga_horaria"],
        data["periodo_meses"], data["telefono"]
    ))
    conn.commit()
    conn.close()
    return jsonify({"mensaje":"registrado"})


# LISTAR BENEFICIARIOS
@app.route("/api/beneficiarios", methods=["GET"])
def obtener_beneficiarios():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT ci, nombre, carga_horaria, fecha_inicio, periodo_meses
    FROM beneficiarios
    """)
    rows = cursor.fetchall()
    beneficiarios = []
    for r in rows:
        fecha_inicio = datetime.strptime(r["fecha_inicio"], "%Y-%m-%d")
        fecha_fin = fecha_inicio + timedelta(days=30 * r["periodo_meses"])
        hoy = datetime.now()
        semanas_restantes = int((fecha_fin - hoy).days / 7)
        beneficiarios.append({
            "ci": r["ci"], "nombre": r["nombre"],
            "carga": r["carga_horaria"], "inicio": r["fecha_inicio"],
            "periodo": r["periodo_meses"], "semanas_restantes": semanas_restantes
        })
    conn.close()
    return jsonify(beneficiarios)


# REGISTRAR ACTIVIDAD EN CRONOGRAMA
@app.route("/api/cronograma", methods=["POST"])
def registrar_actividad():
    data = request.json
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO actividades_mensuales (fecha, descripcion) VALUES (?,?)
    """,(data["fecha"], data["descripcion"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# OBTENER CRONOGRAMA DEL MES
@app.route("/api/cronograma/<anio>/<mes>")
def obtener_cronograma(anio, mes):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT * FROM actividades_mensuales
    WHERE strftime('%Y',fecha)=? AND strftime('%m',fecha)=?
    ORDER BY fecha
    """,(anio, mes.zfill(2)))
    data = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(data)


# ACTIVIDAD DE HOY
@app.route("/api/actividad_hoy")
def actividad_hoy():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT descripcion FROM actividades_mensuales WHERE fecha = DATE('now')")
    r = cursor.fetchone()
    conn.close()
    if r:
        return jsonify({"actividad": r["descripcion"]})
    return jsonify({"actividad": "Sin actividad registrada para hoy"})


# GUARDAR ASISTENCIA
@app.route("/api/registrar_asistencia", methods=["POST"])
def registrar_asistencia():
    registros = request.json
    conn = get_connection()
    cursor = conn.cursor()
    for r in registros:
        ci = r["ci"]
        fecha = r["fecha"]
        dia = r["dia"]
        cursor.execute("SELECT id FROM asistencias WHERE ci=? AND fecha=? AND dia=?", (ci, fecha, dia))
        if cursor.fetchone():
            continue
        hora_inicio = datetime.strptime(r["hora_ingreso"], "%H:%M")
        hora_fin    = datetime.strptime(r["hora_salida"],  "%H:%M")
        horas = (hora_fin - hora_inicio).seconds / 3600
        cursor.execute("""
        INSERT INTO asistencias (ci,fecha,dia,hora_ingreso,hora_salida,horas)
        VALUES (?,?,?,?,?,?)
        """,(ci, fecha, dia, r["hora_ingreso"], r["hora_salida"], horas))
    conn.commit()
    conn.close()
    return jsonify({"mensaje": "ok"})


# LISTA DE ACTIVIDADES (para selector de asistencia)
@app.route("/api/cronograma_lista")
def cronograma_lista():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT fecha, descripcion FROM actividades_mensuales ORDER BY fecha DESC")
    data = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(data)


# HISTORIAL DE ASISTENCIAS POR FECHA
@app.route("/api/asistencia_historial/<fecha>")
def asistencia_historial(fecha):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT b.ci, b.nombre, b.carga_horaria, a.fecha,
           a.hora_ingreso, a.hora_salida, a.horas
    FROM asistencias a
    JOIN beneficiarios b ON a.ci = b.ci
    WHERE a.fecha = ?
    """,(fecha,))
    datos = cursor.fetchall()
    conn.close()
    return jsonify([{
        "ci": d[0], "nombre": d[1], "carga": d[2],
        "fecha": d[3], "hora_ingreso": d[4], "hora_salida": d[5], "horas": d[6]
    } for d in datos])


# DATOS DE UN BENEFICIARIO
@app.route("/api/beneficiario/<ci>")
def obtener_beneficiario(ci):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT ci,nombre,telefono,cite,nudecud,tipo_codigo,fecha_inicio,carga_horaria,periodo_meses
    FROM beneficiarios WHERE ci=?
    """,(ci,))
    b = cursor.fetchone()
    conn.close()
    if not b:
        return jsonify({})
    return jsonify({
        "ci": b[0], "nombre": b[1], "telefono": b[2], "cite": b[3],
        "codigo": b[4], "tipo": b[5], "fecha_inicio": b[6],
        "carga": b[7], "periodo": b[8]
    })


# HISTORIAL DE ASISTENCIAS DEL BENEFICIARIO
@app.route("/api/historial_beneficiario/<ci>")
def historial_beneficiario(ci):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT a.fecha, c.descripcion
    FROM asistencias a
    JOIN actividades_mensuales c ON a.fecha = c.fecha
    WHERE a.ci = ?
    ORDER BY a.fecha
    """,(ci,))
    datos = cursor.fetchall()
    conn.close()
    return jsonify([{"fecha": d[0], "actividad": d[1]} for d in datos])


# ================================================================
# HELPERS COMPARTIDOS PARA GENERACIÓN DE INFORMES
# ================================================================

MESES_ES = ["enero","febrero","marzo","abril","mayo","junio",
            "julio","agosto","septiembre","octubre","noviembre","diciembre"]
DIAS_ES  = ["lunes","martes","miércoles","jueves","viernes","sábado","domingo"]


def fmt_fecha_corta(f):
    """'2026-03-21' → '21 de marzo de 2026'"""
    d = datetime.strptime(f, "%Y-%m-%d")
    return f"{d.day} de {MESES_ES[d.month-1]} de {d.year}"


def fmt_fecha_larga(f):
    """'2026-03-21' → 'sábado 21 de marzo de 2026'"""
    d = datetime.strptime(f, "%Y-%m-%d")
    return f"{DIAS_ES[d.weekday()]} {d.day:02d} de {MESES_ES[d.month-1]} de {d.year}"


def fmt_fecha_obj(d):
    """datetime obj → '21 de marzo de 2026'"""
    return f"{d.day} de {MESES_ES[d.month-1]} de {d.year}"


def calcular_fecha_conclusion(fecha_inicio_str, periodo_meses):
    """
    fecha_inicio + periodo_meses. Si el día no existe en el mes destino
    usa el último día disponible (ej. 31-ene + 1 mes = 28/29-feb).
    """
    inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
    anio_fin = inicio.year + (inicio.month - 1 + periodo_meses) // 12
    mes_fin  = (inicio.month - 1 + periodo_meses) % 12 + 1
    dia_fin  = min(inicio.day, monthrange(anio_fin, mes_fin)[1])
    return datetime(anio_fin, mes_fin, dia_fin)


def calificacion_por_porcentaje(porcentaje):
    """Escala del reglamento GAMEA."""
    if porcentaje == 100:  return "Excelente"
    if porcentaje >= 75:   return "Bueno"
    if porcentaje >= 50:   return "Regular"
    if porcentaje >= 25:   return "Suficiente"
    return "Observado o inexistente"


def construir_lista_asistencias(rows):
    """Rows: (fecha TEXT, descripcion TEXT). Devuelve lista para la plantilla."""
    return [
        {
            "fecha":      fmt_fecha_corta(a[0]),
            "actividad":  a[1] if a[1] else "Actividad comunitaria",
            "asistencia": "SI",
            "es_primera": idx == 0,
        }
        for idx, a in enumerate(rows)
    ]


UNIDADES = [
    "", "un", "dos", "tres", "cuatro", "cinco", "seis", "siete", "ocho", "nueve",
    "diez", "once", "doce", "trece", "catorce", "quince", "dieciséis",
    "diecisiete", "dieciocho", "diecinueve", "veinte", "veintiún", "veintidós",
    "veintitrés", "veinticuatro", "veinticinco", "veintiséis", "veintisiete",
    "veintiocho", "veintinueve",
]
DECENAS = [
    "", "", "veinte", "treinta", "cuarenta", "cincuenta",
    "sesenta", "setenta", "ochenta", "noventa",
]

def numero_a_letras(n):
    """Convierte un entero (0-199) a su texto en español."""
    if n == 0:
        return "cero"
    if n < 30:
        return UNIDADES[n]
    if n < 100:
        dec = DECENAS[n // 10]
        uni = n % 10
        return dec if uni == 0 else f"{dec} y {UNIDADES[uni]}"
    if n == 100:
        return "cien"
    if n < 200:
        uni = n - 100
        return "ciento" if uni == 0 else f"ciento {numero_a_letras(uni)}"
    return str(n)   # fallback para valores fuera de rango


def fmt_semanas(fecha_inicio_str, fecha_conclusion_dt):
    """
    Calcula las semanas exactas entre fecha_inicio y fecha_conclusion
    y devuelve el texto con formato '(52) cincuenta y dos semanas'.
    """
    inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
    dias   = (fecha_conclusion_dt - inicio).days
    semanas = dias // 7
    texto  = numero_a_letras(semanas)
    if semanas == 1:
        return f"({semanas}) una semana"
    return f"({semanas}) {texto} semanas"


# ================================================================
# GENERAR INFORME MENSUAL
# ================================================================

@app.route("/api/generar_informe/<ci>/<anio>/<mes>")
def generar_informe(ci, anio, mes):

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT nombre, ci, carga_horaria, fecha_inicio, periodo_meses
    FROM beneficiarios WHERE ci=?
    """,(ci,))
    b = cursor.fetchone()
    if not b:
        conn.close()
        return jsonify({"error": "Beneficiario no encontrado"}), 404

    cursor.execute("""
    SELECT a.fecha, c.descripcion
    FROM asistencias a
    LEFT JOIN actividades_mensuales c ON a.fecha = c.fecha
    WHERE a.ci=?
      AND strftime('%Y', a.fecha)=?
      AND strftime('%m', a.fecha)=?
    ORDER BY a.fecha
    """,(ci, anio, mes.zfill(2)))
    rows = cursor.fetchall()
    conn.close()

    lista_asistencias = construir_lista_asistencias(rows)

    nombre  = b[0]
    carga   = b[2]
    anio_i  = int(anio)
    mes_i   = int(mes)
    mes_txt = MESES_ES[mes_i - 1]

    if rows:
        periodo_literal = (
            f"{fmt_fecha_larga(rows[0][0])} "
            f"hasta el {fmt_fecha_larga(rows[-1][0])}"
        )
    else:
        periodo_literal = "No registrado"

    # Jornadas obligatorias del mes
    _, dias_en_mes = monthrange(anio_i, mes_i)
    if carga == 16:
        jornadas_a_trabajar = sum(
            1 for d in range(1, dias_en_mes + 1)
            if datetime(anio_i, mes_i, d).weekday() in (0, 5)
        )
    else:
        jornadas_a_trabajar = 4

    jornadas_trabajadas = len(lista_asistencias)
    porcentaje = round((jornadas_trabajadas / jornadas_a_trabajar) * 100) \
                 if jornadas_a_trabajar else 0

    context = {
        "nombre":              nombre,
        "ci":                  b[1],
        "carga_horaria":       carga,
        "periodo_meses":       b[4],
        "mes_anio_mayus_bold": f"{mes_txt.upper()} DE {anio}",
        "mes_anio_minus":      f"{mes_txt} de {anio}",
        "nombre_titulo":       nombre.title(),
        "nombre_mayus":        nombre.upper(),
        "fecha_inicio":        fmt_fecha_larga(b[3]),
        "fecha_inicio_normal": datetime.strptime(b[3], "%Y-%m-%d").strftime("%d/%m/%y"),
        "calcular_semanas":    fmt_semanas(b[3], calcular_fecha_conclusion(b[3], b[4])),
        "periodo_literal":     periodo_literal,
        "asistencias":         lista_asistencias,
        "jornadas_a_trabajar": jornadas_a_trabajar,
        "jornadas_trabajadas": jornadas_trabajadas,
        "porcentaje":          porcentaje,
        "calificacion_asistencia": calificacion_por_porcentaje(porcentaje),
        "calificacion_desempeno":  "Bueno",
        "mes":  mes,
        "anio": anio,
    }

    doc = DocxTemplate(os.path.join(BASE_DIR, "templates_word", "INF. MENSUAL PLANTILLA.docx"))
    doc.render(context)
    stream = BytesIO()
    doc.save(stream)
    stream.seek(0)
    return send_file(
        stream, as_attachment=True,
        download_name=f"informe_{ci}_{anio}_{mes}.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


# ================================================================
# GENERAR INFORME DE CONCLUSIÓN
# ================================================================

@app.route("/api/generar_informe_conclusion/<ci>")
def generar_informe_conclusion(ci):
    """
    Genera el informe de conclusión de penitencia.
    Tabla 1: TODAS las asistencias de toda la condena.
    Tabla 2: evaluación con jornadas totales esperadas vs trabajadas.

    Reglas de jornadas totales:
      - Carga 16h → todos los sábados + lunes en [fecha_inicio, fecha_conclusion]
      - Carga 4/6/8h → 4 jornadas × periodo_meses
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT nombre, ci, carga_horaria, fecha_inicio, periodo_meses
    FROM beneficiarios WHERE ci=?
    """,(ci,))
    b = cursor.fetchone()
    if not b:
        conn.close()
        return jsonify({"error": "Beneficiario no encontrado"}), 404

    nombre        = b[0]
    carga         = b[2]
    fecha_inicio  = b[3]   # "YYYY-MM-DD"
    periodo_meses = b[4]

    # Fecha de conclusión calculada
    conclusion_dt = calcular_fecha_conclusion(fecha_inicio, periodo_meses)

    # Todas las asistencias del período completo
    cursor.execute("""
    SELECT a.fecha, c.descripcion
    FROM asistencias a
    LEFT JOIN actividades_mensuales c ON a.fecha = c.fecha
    WHERE a.ci=?
    ORDER BY a.fecha
    """,(ci,))
    rows = cursor.fetchall()
    conn.close()

    lista_asistencias = construir_lista_asistencias(rows)

    # Jornadas totales esperadas durante toda la condena
    if carga == 16:
        # Recorrer cada día del rango y contar sábados (5) y lunes (0)
        inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        total_jornadas = 0
        cur = inicio_dt
        while cur <= conclusion_dt:
            if cur.weekday() in (0, 5):
                total_jornadas += 1
            cur += timedelta(days=1)
        jornadas_a_trabajar = total_jornadas
    else:
        # 4 jornadas por mes de condena
        jornadas_a_trabajar = 4 * periodo_meses

    jornadas_trabajadas = len(lista_asistencias)
    porcentaje = round((jornadas_trabajadas / jornadas_a_trabajar) * 100) \
                 if jornadas_a_trabajar else 0

    # Período completo para el texto del informe
    if rows:
        periodo_literal = (
            f"{fmt_fecha_larga(rows[0][0])} "
            f"hasta el {fmt_fecha_larga(rows[-1][0])}"
        )
    else:
        periodo_literal = "No registrado"

    cal_asistencia = calificacion_por_porcentaje(porcentaje)

    if cal_asistencia == "Excelente":
        observacion = "satisfactoriamente"
    elif cal_asistencia == "Bueno":
        observacion = "de manera satisfactoria"
    elif cal_asistencia == "Regular":
        observacion = "de manera regular"
    elif cal_asistencia == "Suficiente":
        observacion = "de manera suficiente"
    else:
        observacion = "con observaciones"

    context = {
        "nombre":         nombre,
        "ci":             b[1],
        "carga_horaria":  carga,
        "periodo_meses":  periodo_meses,
        "nombre_titulo":  nombre.title(),
        "nombre_mayus":   nombre.upper(),
        "fecha_inicio_lit":     fmt_fecha_corta(fecha_inicio),
        "fecha_conclusion_lit": fmt_fecha_obj(conclusion_dt),
        "fecha_inicio":         fmt_fecha_larga(fecha_inicio),
        "fecha_inicio_normal":  datetime.strptime(fecha_inicio, "%Y-%m-%d").strftime("%d/%m/%y"),
        "fecha_normal":         datetime.now().strftime("%d/%m/%y"),
        "calcular_semanas":     fmt_semanas(fecha_inicio, conclusion_dt),
        "periodo_literal": periodo_literal,
        "asistencias": lista_asistencias,
        "jornadas_a_trabajar":     jornadas_a_trabajar,
        "jornadas_trabajadas":     jornadas_trabajadas,
        "porcentaje":              porcentaje,
        "calificacion_asistencia": cal_asistencia,
        "calificacion_desempeno":  "Bueno",
        "observacion_desempeno":   observacion,
    }

    doc = DocxTemplate(os.path.join(BASE_DIR, "templates_word", "INF. CONCLUSION PLANTILLA.docx"))
    doc.render(context)
    stream = BytesIO()
    doc.save(stream)
    stream.seek(0)
    return send_file(
        stream, as_attachment=True,
        download_name=f"conclusion_{ci}.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


# ================================================================
# GENERAR PLANILLA DE ASISTENCIA (XLSX)
# ================================================================

@app.route("/api/generar_planilla/<ci>/<anio>/<mes>")
def generar_planilla(ci, anio, mes):
    """
    Genera la planilla mensual usando el template FORMATO_PLANILLA.xlsx como base.
    Respeta exactamente el formato original: fuente, bordes, anchos, títulos.
    Columnas B-I: Nombre | C.I. | CUD/NUREJ | Fecha | Ingreso | Actividad | Salida | Firma
    La columna Firma siempre queda vacía (se firma de forma presencial).
    """
    from openpyxl import load_workbook
    from copy import copy

    conn = get_connection()
    cursor = conn.cursor()

    # 1. Datos del beneficiario
    cursor.execute("""
    SELECT nombre, ci, nudecud, tipo_codigo, carga_horaria, fecha_inicio, periodo_meses
    FROM beneficiarios WHERE ci = ?
    """,(ci,))
    b = cursor.fetchone()
    if not b:
        conn.close()
        return jsonify({"error": "Beneficiario no encontrado"}), 404

    nombre        = b[0]
    ci_ben        = b[1]
    nudecud       = b[2]
    tipo_codigo   = b[3]
    carga         = b[4]

    # 2. Asistencias del mes
    cursor.execute("""
    SELECT a.fecha, a.hora_ingreso, a.hora_salida, c.descripcion
    FROM asistencias a
    LEFT JOIN actividades_mensuales c ON a.fecha = c.fecha
    WHERE a.ci = ?
      AND strftime('%Y', a.fecha) = ?
      AND strftime('%m', a.fecha) = ?
    ORDER BY a.fecha, a.hora_ingreso
    """,(ci, anio, mes.zfill(2)))
    asistencias = cursor.fetchall()
    conn.close()

    mes_txt = MESES_ES[int(mes) - 1].capitalize()

    # 3. Cargar el template original como base
    template_path = os.path.join(BASE_DIR, "templates_word", "FORMATO PLANILLA.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # 4. Actualizar títulos
    ws["B4"] = "PLANILLA DE ASISTENCIA DE TRABAJO COMUNITARIO"
    ws["B5"] = "DIRECCIÓN DE ADMINISTRACIÓN Y MEJORA DE LA INFRAESTRUCTURA Y EQUIPAMIENTO EDUCATIVO"
    ws["B6"] = f"GOBIERNO AUTÓNOMO MUNICIPAL DE LA CIUDAD DE EL ALTO — {mes_txt.upper()} DE {anio}"

    # 5. Encabezado CUD o NUREJ según el beneficiario
    ws["D9"] = tipo_codigo

    # 5b. Activar "Ajustar texto" en toda la tabla (encabezados + filas de datos)
    from openpyxl.styles import Alignment as Aln
    for fila_wrap in range(9, ws.max_row + 1):
        for col_wrap in range(2, 10):   # columnas B=2 .. I=9
            c = ws.cell(row=fila_wrap, column=col_wrap)
            a = c.alignment
            c.alignment = Aln(
                horizontal = a.horizontal,
                vertical   = a.vertical or "center",
                wrap_text  = True,
            )

    # 6. Constantes de layout del template
    COL_MAP = {"B": 2, "C": 3, "D": 4, "E": 5,
               "F": 6, "G": 7, "H": 8, "I": 9}
    PRIMERA_FILA   = 10
    N_FILAS_TMPL   = 4    # el template trae 4 filas vacías (10-13)
    alto_ref       = ws.row_dimensions[PRIMERA_FILA].height or 38.0

    def copiar_estilo(src, dst):
        dst.font          = copy(src.font)
        dst.alignment     = copy(src.alignment)
        dst.border        = copy(src.border)
        dst.fill          = copy(src.fill)
        dst.number_format = src.number_format

    # 7. Insertar filas extra si hay más asistencias que las del template
    n_asis = len(asistencias)
    if n_asis > N_FILAS_TMPL:
        ws.insert_rows(PRIMERA_FILA + N_FILAS_TMPL, n_asis - N_FILAS_TMPL)

    # 8. Rellenar cada fila con los datos de asistencia
    for idx, asis in enumerate(asistencias if asistencias else [("", "", "", "")] * N_FILAS_TMPL):
        fila = PRIMERA_FILA + idx

        # Para filas extra: copiar estilo desde la primera fila de datos
        if idx >= N_FILAS_TMPL:
            for col_num in COL_MAP.values():
                copiar_estilo(
                    ws.cell(row=PRIMERA_FILA, column=col_num),
                    ws.cell(row=fila, column=col_num)
                )
            ws.row_dimensions[fila].height = alto_ref

        # Fecha formateada dd/mm/yyyy
        fecha_str = ""
        if asis[0]:
            try:
                fecha_str = datetime.strptime(asis[0], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                fecha_str = asis[0]

        valores = {
            "B": nombre    if asis[0] else "",
            "C": ci_ben    if asis[0] else "",
            "D": nudecud   if asis[0] else "",
            "E": fecha_str,
            "F": asis[1] or "",
            "G": asis[3] or ("Actividad comunitaria" if asis[0] else ""),
            "H": asis[2] or "",
            "I": "",   # Firma — siempre vacío
        }
        for letra, val in valores.items():
            ws.cell(row=fila, column=COL_MAP[letra], value=val)

    # 9. Enviar el archivo
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"planilla_{ci}_{anio}_{mes}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# EJECUTAR APP
if __name__ == "__main__":
    app.run(debug=True)